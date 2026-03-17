"""Запись итоговой Excel-книги из данных листов."""

from io import BytesIO
from typing import Any

import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.chart.marker import DataPoint
    from openpyxl.drawing.colors import ColorChoice, SchemeColor
except ImportError:
    Workbook = None  # type: ignore[misc, assignment]
    dataframe_to_rows = None  # type: ignore[misc, assignment]
    BarChart = None  # type: ignore[misc, assignment]
    PieChart = None  # type: ignore[misc, assignment]
    Reference = None  # type: ignore[misc, assignment]
    DataLabelList = None  # type: ignore[misc, assignment]
    SeriesLabel = None  # type: ignore[misc, assignment]
    GraphicalProperties = None  # type: ignore[misc, assignment]
    DataPoint = None  # type: ignore[misc, assignment]
    ColorChoice = None  # type: ignore[misc, assignment]
    SchemeColor = None  # type: ignore[misc, assignment]

from app.application.exceptions import ReportBuilderError

# Базовый порядок типов для сводных таблиц и диаграмм. Остальные типы из данных добавляются в конец (отсортированные).
_CHART_BASE_TYPE_ORDER = [
    "АКБ", "Общее", "Чермет", "Цветмет", "Кабель", "Эл.лом", "РЗМ",
    "Нержавейка", "Медь", "Алюминий",
]


def _build_chart_type_order(df: pd.DataFrame) -> list[str]:
    """
    Итоговый список типов для таблиц и диаграмм: базовые в фиксированном порядке,
    затем все остальные уникальные типы из df["Тип"], отсортированные по имени.
    """
    base = list(_CHART_BASE_TYPE_ORDER)
    if "Тип" not in df.columns or df.empty:
        return base
    unique = df["Тип"].dropna().astype(str).unique().tolist()
    dynamic = sorted(t for t in unique if t and t not in base)
    return base + dynamic
# Цвета сегментов первой круговой (Excel accent1..accent6, затем по кругу)
_PIE_COLOR_SCHEME = ["accent1", "accent2", "accent3", "accent4", "accent5", "accent6"]
# Высота первой диаграммы в строках; затем +5 пустых строк до заголовка таблицы расходов
_FIRST_CHART_ROW_SPAN = 15
_GAP_ROWS_BEFORE_EXPENSES = 5
# Нижняя граница второго блока: столбчатая диаграмма расходов занимает по высоте больше строк, чем таблица.
# Высота диаграммы в «строках» для расчёта старта третьего блока (Контакты).
_EXPENSES_CHART_ROW_SPAN = 24
# Отступ перед третьим блоком (Контакты по типам)
_GAP_ROWS_BEFORE_CONTACTS = 5


def _excel_escape(s: str) -> str:
    """Экранирование кавычек для строки в формуле Excel (удвоение ")."""
    if not isinstance(s, str):
        s = str(s) if s is not None else ""
    return s.replace('"', '""')


def _add_pie_chart(
    ws: Any,
    start_row: int,
    title: str,
    col_values: int = 2,
    n_rows: int | None = None,
) -> None:
    """
    Добавляет круговую диаграмму справа от таблицы (колонка D) с теми же настройками подписей.
    Если n_rows задано, диаграмма строится только по первой n_rows строкам данных (после заголовка),
    что позволяет исключать типы с количеством 0 из визуализации, не меняя исходную таблицу.
    """
    chart_anchor = f"D{start_row}"
    if n_rows is None:
        n_rows = len(_CHART_BASE_TYPE_ORDER)
    if n_rows <= 0:
        return
    labels_ref = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + n_rows)
    data_ref = Reference(ws, min_col=col_values, min_row=start_row, max_row=start_row + n_rows)
    pie = PieChart()
    pie.title = title
    pie.width = 22
    pie.height = 14
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(labels_ref)
    if DataLabelList is not None:
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showVal = True
        pie.dataLabels.showPercent = True
        pie.dataLabels.showCatName = False
        pie.dataLabels.showSerName = False
        pie.dataLabels.showLegendKey = False
        pie.dataLabels.position = "bestFit"
    ws.add_chart(pie, chart_anchor)


def _add_pie_chart_contacts(
    ws: Any,
    start_row: int,
    title: str,
    col_labels: int,
    col_values: int,
    n_rows: int,
    type_names_in_order: list[str],
    all_types_order: list[str],
) -> None:
    """
    Круговая диаграмма «Количество контактов»: только значение в подписи (без %),
    цвета сегментов как в остальных диаграммах по типам.
    Категории и значения берутся из одного вспомогательного диапазона (col_labels, col_values).
    """
    if n_rows <= 0:
        return
    chart_anchor = f"D{start_row}"
    labels_ref = Reference(ws, min_col=col_labels, min_row=start_row + 1, max_row=start_row + n_rows)
    data_ref = Reference(ws, min_col=col_values, min_row=start_row, max_row=start_row + n_rows)
    pie = PieChart()
    pie.title = title
    pie.width = 22
    pie.height = 14
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(labels_ref)
    if DataLabelList is not None:
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showVal = True
        pie.dataLabels.showPercent = False
        pie.dataLabels.showCatName = False
        pie.dataLabels.showSerName = False
        pie.dataLabels.showLegendKey = False
        pie.dataLabels.position = "bestFit"
    if GraphicalProperties is not None and DataPoint is not None and SchemeColor is not None and ColorChoice is not None:
        data_points = []
        for i, type_name in enumerate(type_names_in_order):
            color_idx = (all_types_order.index(type_name) if type_name in all_types_order else i) % len(_PIE_COLOR_SCHEME)
            scheme_val = _PIE_COLOR_SCHEME[color_idx]
            fill = ColorChoice(schemeClr=SchemeColor(val=scheme_val))
            sp_pr = GraphicalProperties(solidFill=fill)
            data_points.append(DataPoint(idx=i, spPr=sp_pr))
        if data_points:
            pie.series[0].dPt = data_points
    ws.add_chart(pie, chart_anchor)


def _add_bar_chart_expenses(
    ws: Any,
    start_row: int,
    sorted_type_names: list[str],
    all_types_order: list[str] | None = None,
) -> None:
    """
    Горизонтальная столбчатая: 1 серия «Расходы», категории = типы (ось Y).
    Без заголовка диаграммы. Подписи — только число, справа от столбца (outEnd).
    Цвет столбца = позиция типа в all_types_order (или в базовом списке). Легенда справа.
    """
    chart_anchor = f"D{start_row}"
    n = len(sorted_type_names)
    cats_ref = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + n)
    values_ref = Reference(ws, min_col=2, min_row=start_row + 1, max_row=start_row + n)
    bar = BarChart(barDir="bar")
    bar.width = 22
    bar.height = 14
    bar.add_data(values_ref, from_rows=False, titles_from_data=False)
    bar.set_categories(cats_ref)
    if bar.series and SeriesLabel is not None:
        bar.series[0].title = SeriesLabel(v="Расходы")
    if DataLabelList is not None:
        bar.dataLabels = DataLabelList()
        bar.dataLabels.showVal = True
        bar.dataLabels.showCatName = False
        bar.dataLabels.showSerName = False
        bar.dataLabels.showLegendKey = False
        bar.dataLabels.position = "outEnd"
    type_order = all_types_order or _CHART_BASE_TYPE_ORDER
    if GraphicalProperties is not None and DataPoint is not None and SchemeColor is not None and ColorChoice is not None:
        data_points = []
        for i, type_name in enumerate(sorted_type_names):
            color_idx = (type_order.index(type_name) if type_name in type_order else i) % len(_PIE_COLOR_SCHEME)
            scheme_val = _PIE_COLOR_SCHEME[color_idx]
            fill = ColorChoice(schemeClr=SchemeColor(val=scheme_val))
            sp_pr = GraphicalProperties(solidFill=fill)
            data_points.append(DataPoint(idx=i, spPr=sp_pr))
        bar.series[0].dPt = data_points
    ws.add_chart(bar, chart_anchor)


def _add_chart_sheet_analytics(ws: Any, df: pd.DataFrame) -> None:
    """
    Под основной таблицей листа «Диаграмма» добавляет:
    1) первый блок: сводная «Тип | Количество», круговая «Количество объявлений»;
    2) второй блок ниже: сводная «Тип | Расходы по объявлениям», горизонтальная столбчатая «Расходы по объявлениям».
    Список типов формируется динамически: базовый порядок + остальные типы из данных (отсортированные).
    """
    final_types = _build_chart_type_order(df)
    n_types = len(final_types)
    # Фактическое количество строк основной таблицы на листе «Диаграмма»:
    # заголовок в 1-й строке, данные начинаются со 2-й и идут подряд.
    data_rows = len(df)
    last_data_row = 1 + data_rows  # последняя строка данных основной таблицы
    # Аналитический блок размещаем ниже основной таблицы, оставляя одну пустую строку.
    start_row = last_data_row + 1 + 1  # после заголовка, данных и одной пустой строки

    # --- Первый блок: количество объявлений по типам ---
    ws.cell(row=start_row, column=1, value="Тип")
    ws.cell(row=start_row, column=2, value="Количество по полю Тип")
    counts = df["Тип"].value_counts() if "Тип" in df.columns else pd.Series(dtype=int)
    for i, type_name in enumerate(final_types, start=1):
        row = start_row + i
        ws.cell(row=row, column=1, value=type_name)
        ws.cell(row=row, column=2, value=int(counts.get(type_name, 0)))

    # Для диаграммы используем только типы, у которых количество > 0,
    # при этом сама таблица с нулями остаётся без изменений.
    non_zero_types: list[tuple[str, int]] = []
    for type_name in final_types:
        cnt = int(counts.get(type_name, 0))
        if cnt > 0:
            non_zero_types.append((type_name, cnt))

    if non_zero_types:
        # Вспомогательный диапазон для диаграммы: только строки с количеством > 0,
        # чтобы Reference был непрерывным. Размещаем его справа от основной таблицы
        # (колонки F:G), при необходимости эти колонки можно скрыть в дальнейших итерациях.
        helper_col_label = 6  # F
        helper_col_value = 7  # G
        helper_start_row = start_row
        ws.cell(row=helper_start_row, column=helper_col_label, value="Тип")
        ws.cell(row=helper_start_row, column=helper_col_value, value="Количество по полю Тип")
        for i, (type_name, cnt) in enumerate(non_zero_types, start=1):
            row = helper_start_row + i
            ws.cell(row=row, column=helper_col_label, value=type_name)
            ws.cell(row=row, column=helper_col_value, value=cnt)
        _add_pie_chart(
            ws,
            helper_start_row,
            "Количество объявлений",
            col_values=helper_col_value,
            n_rows=len(non_zero_types),
        )

    # --- Второй блок: расходы — макс. 5 пустых строк после первой диаграммы ---
    second_start = start_row + n_types + _FIRST_CHART_ROW_SPAN + _GAP_ROWS_BEFORE_EXPENSES
    ws.cell(row=second_start, column=1, value="Тип")
    ws.cell(row=second_start, column=2, value="Расходы по объявлениям")
    if "ad_spend" in df.columns:
        spending = df.groupby("Тип", as_index=False)["ad_spend"].sum()
        spend_by_type = spending.set_index("Тип")["ad_spend"]
    else:
        spend_by_type = pd.Series(dtype=float)
    # По убыванию расходов: самый большой сверху, самый маленький снизу (таблица, диаграмма и легенда в одном порядке)
    sorted_pairs = sorted(
        [(t, float(spend_by_type.get(t, 0))) for t in final_types],
        key=lambda x: x[1],
        reverse=True,
    )
    sorted_type_names = [p[0] for p in sorted_pairs]
    for i, (type_name, val) in enumerate(sorted_pairs, start=1):
        row = second_start + i
        ws.cell(row=row, column=1, value=type_name)
        ws.cell(row=row, column=2, value=val)
    if BarChart is not None:
        _add_bar_chart_expenses(ws, second_start, sorted_type_names, all_types_order=final_types)

    # --- Третий блок: контакты по типам. Старт от нижней границы второго блока (таблица или диаграмма), не от конца таблицы. ---
    expenses_table_bottom_row = second_start + n_types
    expenses_chart_bottom_row = second_start + _EXPENSES_CHART_ROW_SPAN
    second_block_bottom_row = max(expenses_table_bottom_row, expenses_chart_bottom_row)
    third_start = second_block_bottom_row + 1 + _GAP_ROWS_BEFORE_CONTACTS
    ws.cell(row=third_start, column=1, value="Тип")
    ws.cell(row=third_start, column=2, value="Контакты")
    if "contacts" in df.columns:
        contacts_by_type = df.groupby("Тип", as_index=False)["contacts"].sum()
        contacts_series = contacts_by_type.set_index("Тип")["contacts"]
    else:
        contacts_series = pd.Series(dtype=float)
    for i, type_name in enumerate(final_types, start=1):
        row = third_start + i
        ws.cell(row=row, column=1, value=type_name)
        ws.cell(row=row, column=2, value=int(contacts_series.get(type_name, 0)))
    non_zero_contacts: list[tuple[str, int]] = []
    for type_name in final_types:
        cnt = int(contacts_series.get(type_name, 0))
        if cnt > 0:
            non_zero_contacts.append((type_name, cnt))
    if non_zero_contacts:
        helper_col_label = 6
        helper_col_value = 7
        ws.cell(row=third_start, column=helper_col_label, value="Тип")
        ws.cell(row=third_start, column=helper_col_value, value="Контакты")
        for i, (type_name, cnt) in enumerate(non_zero_contacts, start=1):
            row = third_start + i
            ws.cell(row=row, column=helper_col_label, value=type_name)
            ws.cell(row=row, column=helper_col_value, value=cnt)
        _add_pie_chart_contacts(
            ws,
            third_start,
            "Количество контактов",
            helper_col_label,
            helper_col_value,
            len(non_zero_contacts),
            [t for t, _ in non_zero_contacts],
            final_types,
        )


def build_workbook_bytes(
    sheets_data: dict[str, pd.DataFrame],
    sheet_order: list[str] | None = None,
    layout: dict[str, Any] | None = None,
) -> bytes:
    """
    Собирает Excel-книгу из словаря {имя_листа: DataFrame}.
    sheet_order задаёт порядок листов; layout — порядок колонок и список колонок по листам.
    Для колонок, у которых в DataFrame есть пара «имя_url» с непустым значением,
    записывается формула =HYPERLINK(url, отображаемый_текст).
    """
    if Workbook is None or dataframe_to_rows is None:
        raise ReportBuilderError(message="openpyxl не установлен")
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    order = sheet_order or list(sheets_data.keys())
    for name in order:
        if name not in sheets_data:
            continue
        df = sheets_data[name]
        ws = wb.create_sheet(title=name)
        # Список колонок для вывода (без _url)
        sheets_layout = (layout or {}).get("sheets") or {}
        sheet_conf = sheets_layout.get(name) or {}
        display_cols = sheet_conf.get("columns")
        if display_cols is None:
            display_cols = [c for c in df.columns if not str(c).endswith("_url")]
        # Заголовок
        for c_idx, col in enumerate(display_cols, start=1):
            ws.cell(row=1, column=c_idx, value=col)
        # Строки данных
        for r_idx in range(len(df)):
            row_num = r_idx + 2
            for c_idx, col in enumerate(display_cols, start=1):
                url_col = f"{col}_url"
                val = df[col].iloc[r_idx] if col in df.columns else None
                url_val = df[url_col].iloc[r_idx] if url_col in df.columns else None
                if url_val is not None and str(url_val).strip() and str(url_val) not in ("nan", "None"):
                    url_str = _excel_escape(str(url_val).strip())
                    disp_str = _excel_escape(str(val) if val is not None and pd.notna(val) else "")
                    cell = ws.cell(row=row_num, column=c_idx)
                    cell.value = f'=HYPERLINK("{url_str}","{disp_str}")'
                else:
                    ws.cell(row=row_num, column=c_idx, value=val)

        # Лист "Диаграмма": сводная таблица по типам и круговая диаграмма под основной таблицей
        if name == "Диаграмма" and "Тип" in df.columns and PieChart is not None and Reference is not None:
            _add_chart_sheet_analytics(ws, df)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def workbook_to_bytes(wb: Any) -> bytes:
    """Сохраняет openpyxl Workbook в bytes."""
    if wb is None:
        raise ReportBuilderError(message="Workbook не передан")
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

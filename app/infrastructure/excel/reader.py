"""Чтение исходного Excel-файла в DataFrame."""

import re
from io import BytesIO
from typing import Any

import pandas as pd

from app.application.exceptions import BrokenExcelError, EmptyFileError

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None  # type: ignore[misc, assignment]

# Извлечение из =HYPERLINK("url", "text")
_HYPERLINK_URL_RE = re.compile(r'=HYPERLINK\s*\(\s*"((?:[^"\\]|\\.)*)"', re.IGNORECASE)
_HYPERLINK_TEXT_RE = re.compile(r',\s*"((?:[^"\\]|\\.)*)"\s*\)', re.IGNORECASE)


def _cell_value(cell: Any) -> Any:
    """Значение ячейки; для формулы HYPERLINK — отображаемый текст."""
    val = cell.value
    if isinstance(val, str) and val.strip().upper().startswith("=HYPERLINK("):
        m = _HYPERLINK_TEXT_RE.search(val)
        return m.group(1).strip() if m else val
    return val


def _cell_url(cell: Any) -> str | None:
    """URL из формулы HYPERLINK; иначе None."""
    val = cell.value
    if isinstance(val, str) and val.strip().upper().startswith("=HYPERLINK("):
        m = _HYPERLINK_URL_RE.search(val)
        return m.group(1).strip() if m else None
    return None


def read_sheet_from_bytes(
    content: bytes,
    sheet_name: str,
    header_row: int = 1,
) -> pd.DataFrame:
    """
    Читает указанный лист из Excel-файла (bytes) в pandas DataFrame.
    header_row — 1-based номер строки с заголовками (по умолчанию 1).
    """
    if not content or len(content) == 0:
        raise EmptyFileError()
    if load_workbook is None:
        raise BrokenExcelError(message="openpyxl не установлен")
    try:
        # read_only=False: корректные dimensions при чтении всех столбцов.
        # data_only=False: нужны формулы HYPERLINK для извлечения отображаемого текста.
        wb = load_workbook(read_only=False, data_only=False, filename=BytesIO(content))
    except Exception as e:
        raise BrokenExcelError(
            message="Файл не удаётся прочитать как Excel",
            details={"error": str(e)},
        ) from e
    try:
        if sheet_name not in wb.sheetnames:
            from app.application.exceptions import RequiredSheetMissingError
            raise RequiredSheetMissingError(
                message=f"Лист «{sheet_name}» не найден",
                details={"available_sheets": wb.sheetnames},
            )
        ws = wb[sheet_name]
        # data_only=False: нужен доступ к формулам HYPERLINK для текста и URL
        rows_raw = list(ws.iter_rows(values_only=False))
    finally:
        wb.close()
    if not rows_raw:
        return pd.DataFrame()
    header_idx = max(0, header_row - 1)
    if header_idx >= len(rows_raw):
        return pd.DataFrame()
    header_cells = rows_raw[header_idx]
    headers = [str(_cell_value(c)).strip() if _cell_value(c) is not None else "" for c in header_cells]
    # Строки: значение + URL для колонок 0 и 7 (Номер объявления, Название объявления)
    data = []
    for row in rows_raw[header_idx + 1 :]:
        values = [_cell_value(c) for c in row]
        url0 = _cell_url(row[0]) if len(row) > 0 else None
        url7 = _cell_url(row[7]) if len(row) > 7 else None
        values.append(url0 or "")
        values.append(url7 or "")
        data.append(values)
    headers_ext = headers + ["Номер объявления_url", "Название объявления_url"]
    df = pd.DataFrame(data, columns=headers_ext)
    return df


def get_sheet_names(content: bytes) -> list[str]:
    """Возвращает список имён листов в книге без полной загрузки данных."""
    if not content or len(content) == 0:
        raise EmptyFileError()
    if load_workbook is None:
        raise BrokenExcelError(message="openpyxl не установлен")
    try:
        wb = load_workbook(read_only=True, data_only=True, filename=BytesIO(content))
        names = list(wb.sheetnames)
        wb.close()
        return names
    except Exception as e:
        raise BrokenExcelError(
            message="Файл не удаётся прочитать как Excel",
            details={"error": str(e)},
        ) from e

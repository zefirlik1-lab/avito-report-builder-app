"""Форматирование итоговой Excel-книги по визуальному ТЗ (expected_report)."""

from io import BytesIO
from typing import Any

from app.infrastructure.config.loader import load_workbook_layout

try:
    from openpyxl import load_workbook as openpyxl_load
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:
    openpyxl_load = None  # type: ignore[misc, assignment]
    Alignment = None  # type: ignore[misc, assignment]
    Border = None  # type: ignore[misc, assignment]
    Font = None  # type: ignore[misc, assignment]
    PatternFill = None  # type: ignore[misc, assignment]
    Side = None  # type: ignore[misc, assignment]

# Визуальное ТЗ
_HEADER_FILL = "4F7F3F"
_HEADER_FONT_COLOR = "FFFFFF"
_BORDER_INNER = "D9D9D9"   # светло-серые внутренние границы
_BORDER_OUTER = "BFBFBF"   # внешняя рамка таблицы чуть темнее
_LINK_COLOR = "0563C1"
_TABLE_HEADER_FILL = "DCE6F1"  # голубая шапка таблиц у диаграмм (Тип | Количество)
_DEFAULT_FONT = "Calibri"
_DEFAULT_FONT_SIZE = 11
_MIN_ROW_HEIGHT = 18

# Выравнивание по умолчанию: текст — слева, числа — справа, проценты — центр
# Специфика по листам (имя колонки → left/center/right)
_SHEET_ALIGNMENT: dict[str, dict[str, str]] = {
    "Админ панель": {
        "Кабинет Авито": "left",
        "Название": "left",
        "Область": "left",
        "Город": "left",
        "Адрес": "left",
        "Номер на Авито": "left",
        "Куда переадресация": "left",
        "ID объявления": "center",
    },
    "Диаграмма": {
        "Кабинет Авито": "left",
        "Тип": "left",
        "Название": "left",
        "Адрес": "left",
        "ID объявления": "left",
    },
    "Недельный": {
        "Категория": "center",
        "Адрес": "left",
        "Номер объявления": "center",
        "Тип": "center",
        "Просмотры": "right",
        "Контакты": "right",
        "Конверсия из просмотров в контакты": "center",
        "Написали в чат": "center",
        "Посмотрели телефон": "center",
        "Средняя цена контакта": "right",
        "Расходы на объявления": "right",
    },
    "Сводная": {
        "Неделя": "center",
        "Период": "center",
        "Просмотры": "right",
        "Конверсия из просмотров в контакты": "center",
        "Контакты": "right",
        "Средняя цена контакта": "right",
        "Звонки": "right",
        "Написали в чат": "right",
        "Расходы на объявления": "right",
        "Комментарий": "left",
    },
}
_WRAP_COLUMNS = {"Название", "Адрес", "Комментарий"}
_LINK_COLUMNS = {"Название", "ID объявления"}


def format_workbook(content: bytes, configs_dir: Any = None) -> bytes:
    """
    Применяет оформление по ТЗ: шрифт Calibri 11, зелёная шапка, границы,
    выравнивание, перенос текста, цвет ссылок, автофильтр, freeze panes.
    """
    if openpyxl_load is None or Font is None or Alignment is None:
        return content
    wb = openpyxl_load(BytesIO(content))
    layout = load_workbook_layout(configs_dir)
    sheets_config = layout.get("sheets") or {}

    header_fill = PatternFill(fill_type="solid", start_color=_HEADER_FILL) if PatternFill else None
    header_font = Font(name=_DEFAULT_FONT, size=_DEFAULT_FONT_SIZE, bold=True, color=_HEADER_FONT_COLOR) if Font else None
    default_font = Font(name=_DEFAULT_FONT, size=_DEFAULT_FONT_SIZE) if Font else None
    link_font = Font(name=_DEFAULT_FONT, size=_DEFAULT_FONT_SIZE, color=_LINK_COLOR, underline="none") if Font else None
    table_header_fill = PatternFill(fill_type="solid", start_color=_TABLE_HEADER_FILL) if PatternFill else None
    table_header_font = Font(name=_DEFAULT_FONT, size=_DEFAULT_FONT_SIZE, bold=True, color="000000") if Font else None

    for sheet_name, sheet_cfg in sheets_config.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        columns = sheet_cfg.get("columns") or []
        align_map = _SHEET_ALIGNMENT.get(sheet_name, {})

        max_row = ws.max_row
        max_col = ws.max_column
        if max_row < 1 or max_col < 1:
            continue

        # Заголовок: зелёный фон, белый жирный текст
        for col in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col)
            if header_fill:
                cell.fill = header_fill
            if header_font:
                cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Автофильтр: только для листов, где диапазон таблицы = весь лист.
        # Для листа «Диаграмма» диапазон автофильтра настраивается внутри _format_diagram_sheet,
        # чтобы не захватывать аналитический блок с маленькими таблицами и диаграммами.
        if sheet_name in ("Админ панель", "Недельный") and max_row > 1 and max_col >= 1:
            ws.auto_filter.ref = f"A1:{_col_letter(max_col)}{max_row}"

        if sheet_name == "Диаграмма":
            _format_diagram_sheet(
                ws, columns, align_map, default_font, link_font,
                table_header_fill, table_header_font,
            )
        else:
            # Данные: шрифт, выравнивание, wrap, цвет ссылок; высота строк (auto при wrap)
            for row_idx in range(2, max_row + 1):
                row_has_wrap = False
                for col_idx in range(1, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if default_font:
                        cell.font = default_font
                    col_name = columns[col_idx - 1] if col_idx <= len(columns) else None
                    horz = align_map.get(col_name, "left") if col_name else "left"
                    if col_name is None and _is_numeric_cell(cell.value):
                        horz = "right"
                    wrap = col_name in _WRAP_COLUMNS if col_name else False
                    if wrap:
                        row_has_wrap = True
                    cell.alignment = Alignment(horizontal=horz, vertical="center", wrap_text=wrap)
                    if col_name in _LINK_COLUMNS and link_font:
                        val = cell.value
                        if isinstance(val, str) and val.strip().upper().startswith("=HYPERLINK("):
                            cell.font = link_font
                if row_has_wrap:
                    ws.row_dimensions[row_idx].height = None
                elif (ws.row_dimensions[row_idx].height or 0) < _MIN_ROW_HEIGHT:
                    ws.row_dimensions[row_idx].height = _MIN_ROW_HEIGHT
            _apply_borders_range(ws, 1, max_row, 1, max_col)

        # Ширины колонок
        widths = sheet_cfg.get("column_widths")
        if isinstance(widths, (list, tuple)):
            for col_idx, w in enumerate(widths, start=1):
                if w is not None and isinstance(w, (int, float)):
                    ws.column_dimensions[_col_letter(col_idx)].width = float(w)

        # Freeze panes: 1 строка для Админ панель и Недельный
        if sheet_name in ("Админ панель", "Недельный"):
            ws.freeze_panes = "A2"

        # Форматы чисел
        if sheet_name == "Недельный" and max_row > 1 and max_col >= 11:
            for row_idx in range(2, max_row + 1):
                ws.cell(row=row_idx, column=7).number_format = "0.00%"
                ws.cell(row=row_idx, column=10).number_format = "0.00"  # Средняя цена контакта
                ws.cell(row=row_idx, column=11).number_format = "0.00"   # Расходы на объявления
        if sheet_name == "Сводная" and max_row > 1 and max_col >= 9:
            for row_idx in range(2, max_row + 1):
                ws.cell(row=row_idx, column=4).number_format = "0.00%"
                ws.cell(row=row_idx, column=6).number_format = "0"
                ws.cell(row=row_idx, column=9).number_format = "0"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _format_diagram_sheet(
    ws: Any,
    columns: list,
    align_map: dict,
    default_font: Any,
    link_font: Any,
    table_header_fill: Any,
    table_header_font: Any,
) -> None:
    """Оформление листа Диаграмма: границы только по данным, стиль таблиц Тип|Количество и Тип|Расходы."""
    max_row = ws.max_row
    max_col = min(ws.max_column, len(columns) or 5)
    if max_row < 2:
        return

    # Найти строку начала первой сводной таблицы (Тип | Количество по полю Тип)
    table1_header_row = None
    for r in range(2, max_row + 1):
        a_val = ws.cell(row=r, column=1).value
        b_val = ws.cell(row=r, column=2).value
        if _cell_str(a_val) == "Тип" and _cell_str(b_val) == "Количество по полю Тип":
            table1_header_row = r
            break
    if table1_header_row is None:
        # Нет блоков диаграмм — оформляем весь лист как одну таблицу
        for row_idx in range(2, max_row + 1):
            _style_diagram_data_row(ws, row_idx, max_col, columns, align_map, default_font, link_font)
            if _row_has_wrap(ws, row_idx, max_col, columns):
                ws.row_dimensions[row_idx].height = None
        _apply_borders_range(ws, 1, max_row, 1, max_col)
        return

    last_main_row = table1_header_row - 2  # основная таблица до одной пустой строки перед блоком
    if last_main_row < 1:
        last_main_row = 1
    main_cols = min(5, max_col)

    # Заголовок уже оформлен в общем цикле. Данные основной таблицы
    for row_idx in range(2, last_main_row + 1):
        _style_diagram_data_row(ws, row_idx, main_cols, columns, align_map, default_font, link_font)
        if _row_has_wrap(ws, row_idx, main_cols, columns):
            ws.row_dimensions[row_idx].height = None
        elif (ws.row_dimensions[row_idx].height or 0) < _MIN_ROW_HEIGHT:
            ws.row_dimensions[row_idx].height = _MIN_ROW_HEIGHT

    _apply_borders_range(ws, 1, last_main_row, 1, main_cols)

    # Автофильтр только по основной таблице, без аналитического блока
    ws.auto_filter.ref = f"A1:{_col_letter(main_cols)}{last_main_row}"

    # Таблица 1: Тип | Количество по полю Тип
    n1 = _count_table_data_rows(ws, table1_header_row, 1)
    _format_chart_table(ws, table1_header_row, table1_header_row + n1, table_header_fill, table_header_font, default_font)

    # Таблица 2: Тип | Расходы по объявлениям
    table2_header_row = None
    for r in range(table1_header_row + n1 + 1, max_row + 1):
        a_val = ws.cell(row=r, column=1).value
        b_val = ws.cell(row=r, column=2).value
        if _cell_str(a_val) == "Тип" and "Расходы" in _cell_str(b_val):
            table2_header_row = r
            break
    if table2_header_row is not None:
        n2 = _count_table_data_rows(ws, table2_header_row, 1)
        _format_chart_table(ws, table2_header_row, table2_header_row + n2, table_header_fill, table_header_font, default_font)


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _row_has_wrap(ws: Any, row_idx: int, max_col: int, columns: list) -> bool:
    for col_idx in range(1, max_col + 1):
        col_name = columns[col_idx - 1] if col_idx <= len(columns) else None
        if col_name in _WRAP_COLUMNS:
            return True
    return False


def _style_diagram_data_row(
    ws: Any, row_idx: int, max_col: int, columns: list, align_map: dict,
    default_font: Any, link_font: Any,
) -> None:
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if default_font:
            cell.font = default_font
        col_name = columns[col_idx - 1] if col_idx <= len(columns) else None
        horz = align_map.get(col_name, "left") if col_name else "left"
        if col_name is None and _is_numeric_cell(cell.value):
            horz = "right"
        wrap = col_name in _WRAP_COLUMNS if col_name else False
        cell.alignment = Alignment(horizontal=horz, vertical="center", wrap_text=wrap)
        if col_name in _LINK_COLUMNS and link_font:
            val = cell.value
            if isinstance(val, str) and val.strip().upper().startswith("=HYPERLINK("):
                cell.font = link_font


def _count_table_data_rows(ws: Any, header_row: int, col: int) -> int:
    n = 0
    for r in range(header_row + 1, header_row + 25):
        if ws.cell(row=r, column=col).value is None and ws.cell(row=r, column=col + 1).value is None:
            break
        n += 1
    return n


# Высота строки заголовка маленьких таблиц на Диаграмме (2 строки текста)
_CHART_TABLE_HEADER_ROW_HEIGHT = 32


def _format_chart_table(
    ws: Any, start_row: int, end_row: int,
    table_header_fill: Any, table_header_font: Any, default_font: Any,
) -> None:
    """Стиль таблицы у диаграммы: голубая шапка, границы D9D9D9/BFBFBF, выравнивание Тип — left, вторая колонка — right."""
    for c in (1, 2):
        cell = ws.cell(row=start_row, column=c)
        if table_header_fill:
            cell.fill = table_header_fill
        if table_header_font:
            cell.font = table_header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[start_row].height = _CHART_TABLE_HEADER_ROW_HEIGHT
    for r in range(start_row + 1, end_row + 1):
        for c in (1, 2):
            cell = ws.cell(row=r, column=c)
            if default_font:
                cell.font = default_font
            cell.alignment = Alignment(horizontal="left" if c == 1 else "right", vertical="center")
    _apply_borders_range(ws, start_row, end_row, 1, 2)


def _apply_borders_range(
    ws: Any,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
    inner_color: str = _BORDER_INNER,
    outer_color: str = _BORDER_OUTER,
) -> None:
    """Границы диапазона: внешняя рамка outer_color, внутренние линии inner_color."""
    if Border is None or Side is None:
        return
    thin_inner = Side(border_style="thin", color=inner_color)
    thin_outer = Side(border_style="thin", color=outer_color)
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            left = thin_outer if c == min_col else thin_inner
            right = thin_outer if c == max_col else thin_inner
            top = thin_outer if r == min_row else thin_inner
            bottom = thin_outer if r == max_row else thin_inner
            ws.cell(row=r, column=c).border = Border(left=left, right=right, top=top, bottom=bottom)


def _is_numeric_cell(value: Any) -> bool:
    """Ячейка с числом для выравнивания по правому краю."""
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str):
        s = value.strip().replace(",", ".").lstrip("-")
        return s.replace(".", "", 1).isdigit() if s else False
    return False


def _col_letter(n: int) -> str:
    """Номер колонки 1-based → буква (1=A, 26=Z, 27=AA)."""
    result = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result or "A"

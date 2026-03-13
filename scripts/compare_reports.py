#!/usr/bin/env python3
"""Сравнение output_report.xlsx и expected_report.xlsx: листы, колонки, данные, форматирование."""
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
OUTPUT = PROJECT_ROOT / "data" / "output_report.xlsx"
EXPECTED = PROJECT_ROOT / "data" / "expected_report.xlsx"

def inspect_workbook(path: Path, name: str) -> dict:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=False, data_only=True)
    info = {
        "path": str(path),
        "name": name,
        "sheet_names": wb.sheetnames,
        "sheets": {},
    }
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        headers = list(rows[0]) if rows else []
        data_rows = len(rows) - 1 if rows else 0
        # Sample: first 2 data rows (as list of tuples)
        sample = [tuple(r) for r in rows[1:3]] if len(rows) > 1 else []
        # Column widths if available
        widths = []
        for c in range(1, len(headers) + 1):
            try:
                col_letter = ws.cell(row=1, column=c).column_letter
                w = ws.column_dimensions[col_letter].width if col_letter in ws.column_dimensions else None
                widths.append(w)
            except Exception:
                widths.append(None)
        info["sheets"][sheet_name] = {
            "headers": headers,
            "num_columns": len(headers),
            "num_data_rows": data_rows,
            "sample_rows": sample,
            "column_widths": widths,
        }
        # Check first cell font (bold for header)
        try:
            first_cell = ws.cell(row=1, column=1)
            info["sheets"][sheet_name]["header_bold"] = getattr(first_cell, "font", None) and getattr(first_cell.font, "bold", None)
        except Exception:
            info["sheets"][sheet_name]["header_bold"] = None
    wb.close()
    return info


def main():
    if not OUTPUT.exists():
        print(f"Missing: {OUTPUT}", file=sys.stderr)
        sys.exit(1)
    if not EXPECTED.exists():
        print(f"Missing: {EXPECTED}", file=sys.stderr)
        sys.exit(1)

    out = inspect_workbook(OUTPUT, "output_report")
    exp = inspect_workbook(EXPECTED, "expected_report")

    print("=" * 60)
    print("1. ЛИСТЫ И ПОРЯДОК")
    print("=" * 60)
    print("output_report sheets:", out["sheet_names"])
    print("expected_report sheets:", exp["sheet_names"])
    print("Order match:", out["sheet_names"] == exp["sheet_names"])
    print()

    print("=" * 60)
    print("2. ПО КАЖДОМУ ЛИСТУ: КОЛОНКИ, СТРОКИ, ПРИМЕР ДАННЫХ")
    print("=" * 60)
    all_sheets = sorted(set(out["sheet_names"]) | set(exp["sheet_names"]))
    for sh in all_sheets:
        print(f"\n--- Лист: {sh} ---")
        o = out["sheets"].get(sh)
        e = exp["sheets"].get(sh)
        if not o:
            print("  [только в expected]")
            print("  headers:", e["headers"])
            print("  data rows:", e["num_data_rows"])
            print("  sample row1:", e["sample_rows"][0] if e["sample_rows"] else None)
            continue
        if not e:
            print("  [только в output]")
            print("  headers:", o["headers"])
            print("  data rows:", o["num_data_rows"])
            continue
        print("  output  headers:", o["headers"])
        print("  expected headers:", e["headers"])
        print("  headers match:", o["headers"] == e["headers"])
        print("  output  data rows:", o["num_data_rows"], "| expected:", e["num_data_rows"])
        if o["sample_rows"]:
            print("  output  sample row1:", o["sample_rows"][0][:5], "...")
        if e["sample_rows"]:
            print("  expected sample row1:", e["sample_rows"][0][:5], "...")
        if o["headers"] == e["headers"] and o["sample_rows"] and e["sample_rows"]:
            for i, (vo, ve) in enumerate(zip(o["sample_rows"][0], e["sample_rows"][0])):
                if vo != ve:
                    print(f"  DIFF col {i} ({o['headers'][i]}): out={repr(vo)[:50]} | exp={repr(ve)[:50]}")
    print()

    print("=" * 60)
    print("3. ФОРМАТИРОВАНИЕ (ширины колонок, жирная шапка)")
    print("=" * 60)
    for sh in all_sheets:
        o = out["sheets"].get(sh)
        e = exp["sheets"].get(sh)
        if not o or not e:
            continue
        print(f"\n{sh}:")
        print("  output  widths:", [round(x, 1) if x else None for x in (o.get("column_widths") or [])[:8]], "...")
        print("  expected widths:", [round(x, 1) if x else None for x in (e.get("column_widths") or [])[:8]], "...")
        print("  output header_bold:", o.get("header_bold"), "| expected:", e.get("header_bold"))


if __name__ == "__main__":
    main()

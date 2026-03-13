#!/usr/bin/env python3
"""Диагностика: колонки и shape сразу после чтения Excel, до нормализации."""
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

INPUT_FILE = PROJECT_ROOT / "data" / "input_example.xlsx"

def main():
    from openpyxl import load_workbook
    from io import BytesIO

    content = INPUT_FILE.read_bytes()
    wb = load_workbook(read_only=True, data_only=True, filename=BytesIO(content))
    ws = wb["Sheet1"]
    data = list(ws.iter_rows(values_only=True))
    wb.close()

    header_idx = 0
    header_row = data[header_idx]
    headers = [str(c).strip() if c is not None else "" for c in header_row]
    rows = data[header_idx + 1:]

    print("1. input_schema.yaml (содержимое):")
    print((PROJECT_ROOT / "app" / "configs" / "input_schema.yaml").read_text(encoding="utf-8"))

    print("2. Длина первой строки (header_row) из openpyxl:", len(header_row))
    print("3. Список колонок сразу после чтения (до нормализации):", headers)
    print("4. Количество колонок (len(headers)):", len(headers))

    import pandas as pd
    df = pd.DataFrame(rows, columns=headers)
    print("5. Shape DataFrame сразу после pd.DataFrame(rows, columns=headers):", df.shape)
    print("6. df.columns.tolist():", df.columns.tolist())

    # Проверка dimensions в openpyxl read_only
    wb2 = load_workbook(read_only=True, data_only=True, filename=BytesIO(content))
    ws2 = wb2["Sheet1"]
    print("7. Worksheet dimensions (read_only): min_row=%s max_row=%s min_col=%s max_col=%s" % (
        getattr(ws2, 'min_row', '?'),
        getattr(ws2, 'max_row', '?'),
        getattr(ws2, 'min_col', '?'),
        getattr(ws2, 'max_col', '?'),
    ))
    wb2.close()

if __name__ == "__main__":
    main()
